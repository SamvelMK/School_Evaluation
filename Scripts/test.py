# Imports 
from sqlalchemy import String, Integer, Float, Boolean, Column, and_, ForeignKey
from connection import Connection
from datetime import datetime, time, date
import time
import pytz
from pytz import timezone
import pandas as pd
import os
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
import openpyxl

from connection import Connection
from user import UserData
from tables import User, Tests, TestsQuestions

class TestData(UserData):
        
    def __init__(self, connection_string):
        
        # First we create a session with the DB.
        self.session = Connection.connection(connection_string)

        # The incomplete cases refer to the number of registered participants 
        # who only took the pre-test but not the post-test.
        # The post_test attempts refer to the number of times the post test was taken.
        
        self.pre = pd.DataFrame()
        self.post = pd.DataFrame()
        self.dropouts = pd.DataFrame()
        self.post_test_attempts = 0
        self.incomplete_cases = 0
    
    def query(self, grade, region_names, trial = None):

        """ 
            If the trial argument is set to be True then the databased is queried with the dates of the trial of the e-learning module. 
            Otherwise, the start date of the date filter is set on the current date (i.e., the date of the query) and the end date is set
            to -1 years. 
        """
        
        if trial:
            start_date =  "2019-06-06 11:00:00"
            end_date = "2019-07-31 00:00:00"
        else:
            start_date = datetime.now()
            end_date = start_date + pd.DateOffset(years=-1)

        self.test = pd.read_sql(self.session.query(User, Tests, TestsQuestions).with_labels()\
                         .filter(and_(Tests.date.between(start_date, end_date), User.grade == str(grade)))\
                                .join(Tests, Tests.user_id == User.id)\
                                .join(TestsQuestions, TestsQuestions.id == Tests.test_id)\
                                .statement, self.session.bind)

        # Changing the time from UCT to US/Central: The "Created_at" and "Updated_at" objects are stored as UNIX time
        # stamp and in UTC timezone. However, other date objects (e.g., in pre_tests_res table) in other tables are 
        # stored as datetime objects and in US Central time zone (based on the location of the IP). 
        # Therefore, 1) we need to change the UNIX int obj to py datetime object; 
        # then 2) change the UTC time zone to US/Central.
        
        # utc = timezone('UTC')
        # central = timezone('US/Central')
        
        self.test['user_created_at'] = self.test['user_created_at'].apply(lambda x: datetime.fromtimestamp(x).strftime("%Y-%m-%d %H:%M:%S"))
                                                               
        self.test['user_updated_at'] = self.test['user_updated_at'].apply(lambda x: datetime.fromtimestamp(x).strftime("%Y-%m-%d %H:%M:%S"))

        # Since the date time is in US Texas timezone I need to convrt it to UCT. 
        local = pytz.timezone ("America/Phoenix")
        self.test['pre_tests_res_date'] = self.test['pre_tests_res_date'].apply(lambda x: local.localize(datetime.strptime(str(x), "%Y-%m-%d %H:%M:%S"), is_dst=None)\
                                                                                    .astimezone(pytz.utc).replace(tzinfo=None))
                                                               
        
        self.test['region_name'] = self.test['user_region'].apply(lambda x: region_names[x])
    
    def test_extract(self, test_dictionary, attitude_dictionary):
        self.knwl_items = test_dictionary
        self.attd_items = attitude_dictionary
        
        # I had to do this because the same person took tests of different grades. :(
        # *knwl does not refer to only knowledge test. in the below assigned variable it inlcudes all test items (including)
        # the attitude items.
        
        
        knwl_flat = test_dictionary.values()
        knwl_flat = [y for x in knwl_flat for y in x]
        
        self.test = self.test[self.test['pre_tests_res_test_id'].isin(knwl_flat)]

        temp_pre = []
        temp_post = []
        temp_dropouts = []
        attempts = 0
        for i in self.test.user_id.unique():
            temp = self.test.loc[self.test['pre_tests_res_user_id'] == i]
            d = temp['pre_tests_res_date'].unique() # would be good to sort it just in case

            if len(d) > 1:
                temp_pre.append(self.test.loc[self.test['pre_tests_res_date'] == d[0]])
                temp_post.append(self.test.loc[self.test['pre_tests_res_date'] == d[1]])
                attempts += len(d) - 2
            elif len(d) == 1:
                temp_dropouts.append(self.test.loc[self.test['pre_tests_res_date'] == d[0]])
                self.incomplete_cases += 1
        
        self.post_test_attempts = attempts  
        self.pre = pd.concat(temp_pre)
        self.post = pd.concat(temp_post)
        self.dropouts = pd.concat(temp_dropouts)
    
    def clean(self):

        def test_clean(test):
            
            """ This function drops the duplicates from the dataframe and returns a dataframe"""
            
            test.drop(['pre_tests_res_user_id'], axis = 1, inplace= True)
            test = test.loc[:,~test.columns.duplicated()]

            temp_test = []
            for i in test.user_id.unique():
                temp = test.loc[test['user_id'] == i]
                temp = temp.drop_duplicates('pre_tests_res_test_id')
                temp_test.append(temp)
                
            test = pd.concat(temp_test)
        
            return test
        
        def answer_check(test, items):
            
            """This function checks the correct answers and returns a dataframe with a new column for correct answers."""
            
            test['answers_checked'] = (test['pre_tests_res_answer'] == test['tests_question_right_answers']).astype(int)
            test.reset_index(inplace = True)
            test.drop('index', axis =1, inplace = True)
            
            for i in items['attitudes']:
                test.at[test.index[test['pre_tests_res_test_id'] == i], 'answers_checked']\
                    = test.iloc[test.index[test['pre_tests_res_test_id'] == i]]['pre_tests_res_answer'].astype(int)
            
            return test
        
        def attach_item_labels(test, items):
            l = []
            for y in test['pre_tests_res_test_id']:
                for i in items:
                    if y in items[i]:
                        l.append(i)
            test['topic'] = l
                        
            return test
        
        # Knowledge items
        pre = answer_check(test_clean(self.pre), self.knwl_items)
        post = answer_check(test_clean(self.post), self.knwl_items)

        self.pre = attach_item_labels(pre, self.knwl_items)
        self.post = attach_item_labels(post, self.knwl_items)
        
        # Attitude items
        pre_attd = self.pre[self.pre['topic'] == 'attitudes']
        post_attd = self.post[self.post['topic'] == 'attitudes']

        pre_attd['answers_checked'] = pre_attd['answers_checked'].apply(lambda x: 11 - x) # This reverse codes the attitude scale
        post_attd['answers_checked'] = post_attd['answers_checked'].apply(lambda x: 11 - x) # This reverse codes the attitude scale

        self.pre_attd = attach_item_labels(pre_attd, self.attd_items)
        self.post_attd = attach_item_labels(post_attd, self.attd_items)
        

    def test_info(self):
        
        self.test_takers = len(self.test['user_id'].unique())
        self.completed_cases = len(self.pre['user_id'].unique())
        
        def correct_answers(test):
            
            correct_answers = (test[~test['pre_tests_res_test_id'].isin(self.knwl_items['attitudes'])]\
                                .answers_checked.sum())*100/len(test.answers_checked)
            
            return correct_answers
        
        def correct_by_sex(test):
            
            correct_by_sex = test[~test['pre_tests_res_test_id'].isin(self.knwl_items['attitudes'])]\
                    .groupby('user_sex')['answers_checked'].sum()*100/test[~test['pre_tests_res_test_id'].isin(self.knwl_items['attitudes'])]\
                    .groupby('user_sex')['answers_checked'].count() 
            
            return correct_by_sex
        
        self.correct_answers_pre = correct_answers(self.pre)
        self.correct_answers_post = correct_answers(self.post)
        
        self.correct_by_sex_pre = correct_by_sex(self.pre)
        self.correct_by_sex_post = correct_by_sex(self.post)
        curent_date = date.today().strftime("%B %d, %Y")       
        self.info = """ 
        
        ************************************************************************************************
        ****                                Test INFO                                               ****
        ************************************************************************************************
        **** Data acquisition date: {}                                                 ****
        **** Overall results:                                                                       ****
        ****                                                                                        ****
        **** \t Number of overall test takers: {}                                                  ****
        **** \t Number of people who took only the pre test: {}                                    ****
        **** \t Numbr of students who completed both the pre and the post tests: {}                ****
        ****                                                                                        ****
        **** Pre Test:                                                                              ****
        ****                                                                                        ****
        **** \t Proportion of correct answers: {}%                                              ****
        **** \t Proportion of overall correct answers on knwoledge questions per sex:              **** 
        **** \t Female: {}%                                                                     ****
        **** \t Male: {}%                                                                       ****
        ****                                                                                        ****
        **** Post Test :                                                                            ****
        ****                                                                                        ****
        **** \t Proportion of correct answers: {}%                                              ****
        **** \t Proportion of overall correct answers on knwoledge questions per sex:              ****
        **** \t Female: {}%                                                                     ****
        **** \t Male: {}%                                                                       ****
        ************************************************************************************************
        **** Note: The number of pre and post test takers may differ.                               ****
        ************************************************************************************************
        
        """.format(curent_date, self.test_takers, len(self.dropouts['user_id'].unique()), self.completed_cases, 
                   self.correct_answers_pre.round(2), self.correct_by_sex_pre[0].round(2), self.correct_by_sex_pre[1].round(2),
                   self.correct_answers_post.round(2), self.correct_by_sex_post[0].round(2), self.correct_by_sex_post[1].round(2))

        return self.info
    
    
    #### Here you can find methods for simple aggregate stats
    
    def aggr_corr_prop(self, test, items, test_name, *args):
        
        """ DESCRIBE THE METHOD"""
        
        df = test[~test['pre_tests_res_test_id'].isin(items)]
        df_proportions = pd.DataFrame((df.groupby([i for i in args])['answers_checked'].sum()*100\
                                        /df.groupby([i for i in args])['answers_checked'].count()).round(2))
        
        df_counts = pd.DataFrame((df.groupby([i for i in args])['pre_tests_res_test_id'].nunique()))
        
        new = pd.merge(df_proportions, df_counts, on = [i for i in args])
        
        new.rename(columns={'answers_checked':'Proportion_Correct: {}'.format(test_name),
                          'pre_tests_res_test_id':'Numbr of Items Answered: {}'.format(test_name)},
                        inplace = True)
        
        return new
    
    def aggr_corr_mean(self, test, items, test_name, *args):
        
        """ DESCRIBE THE METHOD"""
        
        df = test[~test['pre_tests_res_test_id'].isin(items)]
        df_mean = pd.DataFrame((df.groupby([i for i in args])['answers_checked'].mean()).round(2))
        df_sd = pd.DataFrame((df.groupby([i for i in args])['answers_checked'].std()).round(2))
        
        new = pd.merge(df_mean, df_sd, on = [i for i in args])
        
        new.rename(columns={'answers_checked_x':'Mean: {}'.format(test_name),
                          'answers_checked_y':'SD: {}'.format(test_name)},
                        inplace = True)
                   
        return new
    
    def combine(self, pre_test, post_test, merge_on, names):
        
        """ DESCRIBE THE METHOD"""
                
        test_aggr_region = pd.merge(pre_test, post_test, on = merge_on)
        test_aggr_region.index.set_names(names = [i for i in names], inplace = True)
        
        return test_aggr_region
    
    ### Methods for knowledge questions
    
    def knwl_aggregate_correct_region(self):

        pre = self.aggr_corr_prop(self.pre, self.knwl_items['attitudes'], 'pre', 'region_name')
        post = self.aggr_corr_prop(self.post, self.knwl_items['attitudes'], 'post', 'region_name') 
        self.test_aggr_region = self.combine(pre, post, ['region_name'], ['Region'])
    
    def knwl_aggregate_correct_reg_sex(self):
        
        pre = self.aggr_corr_prop(self.pre, self.knwl_items['attitudes'], 'pre', 'region_name', 'user_sex')
        post = self.aggr_corr_prop(self.post, self.knwl_items['attitudes'], 'post', 'region_name', 'user_sex')  
        self.test_aggr_region_sex = self.combine(pre, post, ['region_name', 'user_sex'], ['Region', 'Sex'])
    
    def knwl_aggregate_thematic(self):
        
        pre = self.aggr_corr_prop(self.pre, self.knwl_items['attitudes'], 'pre', 'topic')
        post = self.aggr_corr_prop(self.post, self.knwl_items['attitudes'], 'post', 'topic')
        self.test_aggr_tematic = self.combine(pre, post, ['topic'], ['Topic'])
    
    def knwl_aggregate_thematic_reg(self):
        
        pre = self.aggr_corr_prop(self.pre, self.knwl_items['attitudes'], 'pre', 'region_name', 'topic')
        post = self.aggr_corr_prop(self.post,  self.knwl_items['attitudes'], 'post', 'region_name', 'topic')
        
        self.test_aggr_theme_reg = self.combine(pre, post, ['region_name', 'topic'], ['Region', 'Topic'])
    
    def knwl_aggregate_thematic_sex(self):
        
        pre = self.aggr_corr_prop(self.pre, self.knwl_items['attitudes'], 'pre', 'topic', 'user_sex')
        post = self.aggr_corr_prop(self.post, self.knwl_items['attitudes'], 'post', 'topic', 'user_sex')
        
        self.test_aggr_theme_sex = self.combine(pre, post, ['topic', 'user_sex'], ['Topic','Sex'])
    
    def knwl_aggregate_thematic_reg_sex(self):
        
        pre = self.aggr_corr_prop(self.pre, self.knwl_items['attitudes'], 'pre', 'region_name', 'topic', 'user_sex')
        post = self.aggr_corr_prop(self.post, self.knwl_items['attitudes'], 'post', 'region_name', 'topic', 'user_sex')

        self.test_aggr_theme_reg_sex = self.combine(pre, post, ['region_name', 'topic', 'user_sex'], ['Region', 'Topic', 'Sex'])
    
    ### Methods for attitude items
    
    def attd_mean_items_theme(self):
        
        pre_mean = self.aggr_corr_mean(self.pre_attd, self.attd_items, 'pre', 'topic')
        post_mean = self.aggr_corr_mean(self.post_attd, self.attd_items, 'post', 'topic')
        self.attd_mean_theme = self.combine(pre_mean, post_mean, 'topic', ['Topic'])
    
    def attd_mean_items_theme_sex(self):
        
        pre_mean = self.aggr_corr_mean(self.pre_attd, self.attd_items, 'pre', 'topic', 'user_sex')
        post_mean = self.aggr_corr_mean(self.post_attd, self.attd_items, 'post', 'topic', 'user_sex')
        
        self.attd_mean_theme_sex = self.combine(pre_mean, post_mean, ['topic', 'user_sex'], ['Topic', 'Sex'])
    
    def attd_mean_items_theme_reg(self):
        
        pre_mean = self.aggr_corr_mean(self.pre_attd, self.attd_items, 'pre', 'region_name', 'topic')
        post_mean = self.aggr_corr_mean(self.post_attd, self.attd_items, 'post', 'region_name', 'topic')
        
        self.attd_mean_theme_reg = self.combine(pre_mean, post_mean, ['region_name', 'topic'], ['Region', 'Topic'])
    
    def attd_mean_items_theme_reg_sex(self):
        
        pre_mean = self.aggr_corr_mean(self.pre_attd, self.attd_items, 'pre', 'region_name', 'topic', 'user_sex')
        post_mean = self.aggr_corr_mean(self.post_attd, self.attd_items, 'post', 'region_name', 'topic', 'user_sex')
        
        self.attd_mean_theme_reg_sex = self.combine(pre_mean, post_mean, ['region_name', 'topic', 'user_sex'], ['Region', 'Topic', 'Sex'])
    
    @staticmethod
    def write_to_xls(directory, filename, trial = None, **kwargs):
        
        current_dir = os.getcwd()

        if trial:
            now = datetime.now()
            year = now.year
        else:
            year = now.year - 1

        if not os.path.exists('Data'):
            os.mkdir('Data')
        os.chdir('Data')
         
        if not os.path.exists('{}'.format(directory)):
            os.mkdir('{}'.format(directory))
        os.chdir('{}'.format(directory)) 
        
        if not os.path.exists('{}_{}'.format(directory, year)):
            os.mkdir('{}_{}'.format(directory, year))
        os.chdir('{}_{}'.format(directory, year))    
            
        if not os.path.isfile('{}_{}.xlsx'.format(filename, year)):
            writer = pd.ExcelWriter('{}_{}.xlsx'.format(filename, year),  engine='xlsxwriter') 
            for i in kwargs:
                kwargs[i].to_excel(writer, sheet_name='{}'.format(i))   
            writer.save()
        else:
            wb = openpyxl.load_workbook('{}_{}.xlsx'.format(filename, year))
            for i in kwargs:             
                writer = pd.ExcelWriter('{}_{}.xlsx'.format(filename, year), engine='openpyxl') 
                writer.book = wb
                sheets = wb.sheetnames
                if not i in sheets:
                    kwargs[i].to_excel(writer, sheet_name='{}'.format(i))
                    writer.save()
            writer.close()            
            
        os.chdir(current_dir)